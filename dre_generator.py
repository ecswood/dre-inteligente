import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, DoughnutChart
import json

def generate_dre():
    print("Iniciando geração da DRE...")
    
    # 1. Carregar planilha de lançamentos original
    wb_source = openpyxl.load_workbook('caixahistorico.xlsx', data_only=True)
    ws_source = wb_source.active
    
    # Encontrar a linha de cabeçalho (que contém 'Código')
    header_row_idx = None
    for r_idx, row in enumerate(ws_source.iter_rows(values_only=True), 1):
        if 'Código' in row:
            header_row_idx = r_idx
            break
            
    if header_row_idx is None:
        raise ValueError("Cabeçalho 'Código' não encontrado na planilha de origem!")
        
    print(f"Cabeçalho encontrado na linha {header_row_idx}")
    
    # Extrair os dados da planilha
    source_rows = list(ws_source.iter_rows(values_only=True))
    headers = source_rows[header_row_idx - 1]
    data_rows = source_rows[header_row_idx:] # dados a partir da linha seguinte ao cabeçalho
    
    # 2. Carregar o arquivo de mapeamento JSON
    with open('isp_dre_mapping.json', 'r', encoding='utf-8') as f:
        mapping = json.load(f)
        
    # 3. Criar a nova pasta de trabalho
    wb = openpyxl.Workbook()
    
    # Remover aba padrão vazia
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Criar abas
    ws_lanc = wb.create_sheet(title="Lançamentos")
    ws_depara = wb.create_sheet(title="De_Para")
    ws_dre = wb.create_sheet(title="DRE")
    
    # --- ABA DE_PARA ---
    ws_depara.append(["Plano de Contas ERP", "Categoria DRE"])
    # Formatar cabeçalho
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    border_thin = Border(
        left=Side(style='thin', color='BFBFBF'),
        right=Side(style='thin', color='BFBFBF'),
        top=Side(style='thin', color='BFBFBF'),
        bottom=Side(style='thin', color='BFBFBF')
    )
    
    for col in range(1, 3):
        cell = ws_depara.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_thin
        
    # Escrever mapeamentos no De_Para
    row_count = 2
    for erp_cat, map_info in sorted(mapping.items()):
        dre_cat = map_info["dre_category"]
        if map_info["action"] == "exclude":
            dre_cat = "EXCLUÍDO DA DRE"
        ws_depara.append([erp_cat, dre_cat])
        ws_depara.cell(row=row_count, column=1).border = border_thin
        ws_depara.cell(row=row_count, column=2).border = border_thin
        row_count += 1
        
    # Ajustar largura de colunas no De_Para
    ws_depara.column_dimensions['A'].width = 45
    ws_depara.column_dimensions['B'].width = 45
    
    # --- ABA LANÇAMENTOS ---
    # Cabeçalho da aba lançamentos (adicionamos a coluna K: Categoria DRE)
    ws_lanc.append(list(headers) + ["Categoria DRE"])
    
    # Formatar cabeçalho
    for col in range(1, len(headers) + 2):
        cell = ws_lanc.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_thin
        
    # Escrever dados brutos e adicionar a fórmula do VLOOKUP
    print("Escrevendo lançamentos e aplicando fórmulas PROCV...")
    for idx, row in enumerate(data_rows, 2):
        # Escrever linha de dados (apenas as 10 colunas originais)
        ws_lanc.append(list(row[:10]) + [None])
        
        # Inserir fórmula na coluna K (11)
        # Se o lançamento for da conta "02.09 : Outras Saídas", aplicamos regras de override baseadas em palavras-chave.
        # Caso contrário, usamos o VLOOKUP tradicional na aba De_Para.
        formula = (
            f'=IF(H{idx}="02.09 : Outras Saídas", '
            f'IF(OR('
            f'ISNUMBER(SEARCH("CHEQUE COMPE SICREDI", F{idx})), '
            f'ISNUMBER(SEARCH("aniversario", F{idx})), '
            f'ISNUMBER(SEARCH("churrasco", F{idx})), '
            f'ISNUMBER(SEARCH("bolo", F{idx})), '
            f'ISNUMBER(SEARCH("SUPERMERCADO", F{idx}))'
            f'), "6.1 Despesas Administrativas (Pessoal)", '
            f'IF(OR(ISNUMBER(SEARCH("DLKNET", F{idx})), ISNUMBER(SEARCH("TESTE", F{idx}))), "6.1 Despesas Administrativas (Gerais)", '
            f'IFERROR(VLOOKUP(H{idx}, De_Para!A:B, 2, FALSE), "NÃO MAPEADO"))), '
            f'IFERROR(VLOOKUP(H{idx}, De_Para!A:B, 2, FALSE), "NÃO MAPEADO"))'
        )
        ws_lanc.cell(row=idx, column=11, value=formula)
        
        # Formatações das colunas de Entrada (9) e Saída (10) como moeda
        for col in [9, 10]:
            val_cell = ws_lanc.cell(row=idx, column=col)
            if val_cell.value is not None:
                val_cell.number_format = 'R$ #,##0.00; (R$ #,##0.00); "-"'
                
        # Bordas leves
        for col in range(1, 12):
            ws_lanc.cell(row=idx, column=col).border = border_thin
            
    # Ajustar larguras de coluna na aba Lançamentos
    col_widths = {
        'A': 10, 'B': 20, 'C': 15, 'D': 15, 'E': 15,
        'F': 40, 'G': 12, 'H': 35, 'I': 15, 'J': 15, 'K': 35
    }
    for col_letter, width in col_widths.items():
        ws_lanc.column_dimensions[col_letter].width = width
        
    # --- ABA DRE ---
    # Definindo estilos para a aba DRE
    font_title = Font(name="Segoe UI", size=16, bold=True, color="FFFFFF")
    font_subtitle = Font(name="Segoe UI", size=11, italic=True, color="FFFFFF")
    font_group = Font(name="Segoe UI", size=11, bold=True, color="000000")
    font_subgroup = Font(name="Segoe UI", size=10, color="000000")
    font_total = Font(name="Segoe UI", size=11, bold=True, color="000000")
    
    fill_title = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    fill_subtitle = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
    fill_header = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_total = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill_net_income = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # soft green
    
    border_total = Border(
        top=Side(style='thin', color='000000'),
        bottom=Side(style='double', color='000000')
    )
    
    # 1. Título
    ws_dre.merge_cells("A1:C1")
    ws_dre["A1"] = "DEMONSTRATIVO DE RESULTADOS DO EXERCÍCIO - DRE"
    ws_dre["A1"].font = font_title
    ws_dre["A1"].fill = fill_title
    ws_dre["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_dre.row_dimensions[1].height = 35
    
    ws_dre.merge_cells("A2:C2")
    ws_dre["A2"] = "Provedor de Internet (ISP) - Relatório Gerencial Dinâmico"
    ws_dre["A2"].font = font_subtitle
    ws_dre["A2"].fill = fill_subtitle
    ws_dre["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws_dre.row_dimensions[2].height = 20
    
    # Cabeçalho da Tabela
    ws_dre.cell(row=4, column=1, value="Estrutura do Plano de Contas").font = font_group
    ws_dre.cell(row=4, column=1).fill = fill_header
    ws_dre.cell(row=4, column=1).alignment = Alignment(horizontal="left", vertical="center")
    
    ws_dre.cell(row=4, column=2, value="Valor (R$)").font = font_group
    ws_dre.cell(row=4, column=2).fill = fill_header
    ws_dre.cell(row=4, column=2).alignment = Alignment(horizontal="right", vertical="center")
    
    ws_dre.cell(row=4, column=3, value="% Rec. Bruta").font = font_group
    ws_dre.cell(row=4, column=3).fill = fill_header
    ws_dre.cell(row=4, column=3).alignment = Alignment(horizontal="right", vertical="center")
    ws_dre.row_dimensions[4].height = 25
    
    # Estruturar as linhas da DRE
    # Cada entrada: (Nome da Conta, Fórmula de Valor ou Categoria de Mapeamento, Estilo Tipo, Fórmula % ROL)
    # Estilo Tipos: 'group' (bold), 'subgroup' (indent), 'total' (grey fill + borders), 'net_income' (green fill + bold)
    
    dre_structure = [
        ("1. RECEITA OPERACIONAL BRUTA", "=SUM(B6:B10)", "group", None), # R5
        ("  1.2 Receita de Internet (SVA)", "1.2 Receita de Internet (SVA)", "subgroup", "entradas"), # R6
        ("  1.3 Taxas de Instalação e Adesão", "1.3 Taxas de Instalação e Adesão", "subgroup", "entradas"), # R7
        ("  1.4 Outras Receitas (Equipamentos)", "1.4 Outras Receitas (Equipamentos)", "subgroup", "entradas"), # R8
        ("  1.4 Outras Receitas (Serviços Avulsos)", "1.4 Outras Receitas (Serviços Avulsos)", "subgroup", "entradas"), # R9
        ("  1.4 Outras Receitas (Diversas)", "1.4 Outras Receitas (Diversas)", "subgroup", "entradas"), # R10
        ("  (=) Total Receita Bruta", "=SUM(B6:B10)", "total", "=B11/$B$11"), # R11
        ("2. (-) DEDUÇÕES E TRIBUTOS", "=B14", "group", None), # R12
        ("  2.1 Tributos sobre Serviços (DARF)", "2.1 Tributos sobre Serviços (DARF)", "subgroup", "saidas"), # R13
        ("  (=) Total Deduções", "=B13", "total", "=B14/$B$11"), # R14
        ("3. (=) RECEITA OPERACIONAL LÍQUIDA (ROL)", "=B11-B14", "net_income", "=B15/$B$11"), # R15
        ("4. (-) CUSTOS DOS SERVIÇOS PRESTADOS (CSP)", "=B20", "group", None), # R16
        ("  4.1 Links Dedicados / Trânsito IP", "4.1 Links Dedicados / Trânsito IP", "subgroup", "saidas"), # R17
        ("  4.2 Postes e Aluguel de Infraestrutura", "4.2 Postes e Aluguel de Infraestrutura", "subgroup", "saidas"), # R18
        ("  4.5 Manutenção de Rede / Licenças Técnicas", "4.5 Manutenção de Rede / Licenças Técnicas", "subgroup", "saidas"), # R19
        ("  (=) Total Custos (CSP)", "=SUM(B17:B19)", "total", "=B20/$B$11"), # R20
        ("5. (=) MARGEM DE CONTRIBUIÇÃO / LUCRO BRUTO", "=B15-B20", "net_income", "=B21/$B$11"), # R21
        ("6. (-) DESPESAS OPERACIONAIS (OPEX)", "=B30", "group", None), # R22
        ("  6.1 Despesas Administrativas (Pessoal)", "6.1 Despesas Administrativas (Pessoal)", "subgroup", "saidas"), # R23
        ("  6.1 Despesas Administrativas (Infra)", "6.1 Despesas Administrativas (Infra)", "subgroup", "saidas"), # R24
        ("  6.1 Despesas Administrativas (Gerais)", "6.1 Despesas Administrativas (Gerais)", "subgroup", "saidas"), # R25
        ("  6.3 Despesas Financeiras (Taxas Boleto)", "6.3 Despesas Financeiras (Taxas Boleto)", "subgroup", "saidas"), # R26
        ("  6.3 Despesas Financeiras (Gerais)", "6.3 Despesas Financeiras (Gerais)", "subgroup", "saidas"), # R27
        ("  6.3 Despesas Financeiras (Encargos)", "6.3 Despesas Financeiras (Encargos)", "subgroup", "saidas"), # R28
        ("  6.4 Outras Despesas Operacionais", "6.4 Outras Despesas Operacionais", "subgroup", "saidas"), # R29
        ("  (=) Total Despesas (OPEX)", "=SUM(B23:B29)", "total", "=B30/$B$11"), # R30
        ("7. (=) RESULTADO OPERACIONAL (EBITDA)", "=B21-B30", "net_income", "=B31/$B$11"), # R31
        ("8. (-) AMORTIZAÇÃO E OUTROS", "=B34", "group", None), # R32
        ("  8.1 Amortização de Empréstimos e Financiamentos", "8.1 Amortização de Empréstimos e Financiamentos", "subgroup", "saidas"), # R33
        ("  (=) Total Amortização", "=B33", "total", "=B34/$B$11"), # R34
        ("9. (=) RESULTADO LÍQUIDO DO EXERCÍCIO", "=B31-B34", "net_income", "=B35/$B$11") # R35
    ]
    
    for r_idx, (name, val_source, style_type, pct_formula) in enumerate(dre_structure, 5):
        # 1. Nome da conta
        ws_dre.cell(row=r_idx, column=1, value=name)
        
        # 2. Valor
        if str(val_source).startswith("="):
            # É uma fórmula
            ws_dre.cell(row=r_idx, column=2, value=val_source)
        else:
            # É uma categoria de mapeamento
            if pct_formula == "entradas":
                formula = f'=SUMIFS(Lançamentos!I:I, Lançamentos!K:K, "{val_source}")'
            elif pct_formula == "saidas":
                formula = f'=-SUMIFS(Lançamentos!J:J, Lançamentos!K:K, "{val_source}")'
            else:
                formula = 0
            ws_dre.cell(row=r_idx, column=2, value=formula)
            
        # 3. % Receita Bruta
        if pct_formula not in ["entradas", "saidas"] and pct_formula is not None:
            ws_dre.cell(row=r_idx, column=3, value=pct_formula)
        elif pct_formula in ["entradas", "saidas"]:
            ws_dre.cell(row=r_idx, column=3, value=f'=B{r_idx}/$B$11')
            
        # Formatar células de acordo com o estilo
        cell_name = ws_dre.cell(row=r_idx, column=1)
        cell_val = ws_dre.cell(row=r_idx, column=2)
        cell_pct = ws_dre.cell(row=r_idx, column=3)
        
        # Configurar formatos padrão
        cell_val.number_format = 'R$ #,##0.00; (R$ #,##0.00); "-"'
        cell_pct.number_format = '0.0%'
        
        if style_type == 'group':
            cell_name.font = font_group
            cell_val.font = font_group
            cell_pct.font = font_group
        elif style_type == 'subgroup':
            cell_name.font = font_subgroup
            cell_val.font = font_subgroup
            cell_pct.font = font_subgroup
            # Adicionar recuo visual
            cell_name.alignment = Alignment(indent=1)
        elif style_type == 'total':
            cell_name.font = font_total
            cell_val.font = font_total
            cell_pct.font = font_total
            cell_name.fill = fill_total
            cell_val.fill = fill_total
            cell_pct.fill = fill_total
            cell_name.border = Border(top=Side(style='thin', color='A0A0A0'), bottom=Side(style='thin', color='A0A0A0'))
            cell_val.border = Border(top=Side(style='thin', color='A0A0A0'), bottom=Side(style='thin', color='A0A0A0'))
            cell_pct.border = Border(top=Side(style='thin', color='A0A0A0'), bottom=Side(style='thin', color='A0A0A0'))
        elif style_type == 'net_income':
            cell_name.font = font_total
            cell_val.font = font_total
            cell_pct.font = font_total
            cell_name.fill = fill_net_income
            cell_val.fill = fill_net_income
            cell_pct.fill = fill_net_income
            cell_name.border = border_total
            cell_val.border = border_total
            cell_pct.border = border_total
            
    ws_dre.column_dimensions['A'].width = 50
    ws_dre.column_dimensions['B'].width = 22
    ws_dre.column_dimensions['C'].width = 15
    
    # Habilitar linhas de grade em todas as abas
    ws_lanc.views.sheetView[0].showGridLines = True
    ws_depara.views.sheetView[0].showGridLines = True
    ws_dre.views.sheetView[0].showGridLines = True
    
    # --- ADICIONAR GRÁFICOS NATIVOS DO EXCEL ---
    print("Gerando gráficos nativos do Excel...")
    
    # 1. Gráfico de Custos do ISP (CSP) - Barras Horizontais
    chart_costs = BarChart()
    chart_costs.type = "bar" # horizontal
    chart_costs.title = "Distribuição de Custos ISP (CSP)"
    chart_costs.style = 10
    
    # Referenciar dados de Custos (linhas 17 a 19)
    data_costs = Reference(ws_dre, min_col=2, min_row=17, max_row=19)
    cats_costs = Reference(ws_dre, min_col=1, min_row=17, max_row=19)
    chart_costs.add_data(data_costs, titles_from_data=False)
    chart_costs.set_categories(cats_costs)
    chart_costs.legend = None # Sem legenda pois as barras já têm rótulo
    chart_costs.width = 18
    chart_costs.height = 8.5
    ws_dre.add_chart(chart_costs, "E4")
    
    # 2. Gráfico de Despesas Administrativas e Financeiras (OPEX) - Donut/Rosca
    chart_opex = DoughnutChart()
    chart_opex.title = "Composição de Despesas (OPEX)"
    chart_opex.style = 2  # Estilo clássico limpo
    
    # Referenciar dados de Despesas (linhas 23 a 29)
    data_opex = Reference(ws_dre, min_col=2, min_row=23, max_row=29)
    cats_opex = Reference(ws_dre, min_col=1, min_row=23, max_row=29)
    chart_opex.add_data(data_opex, titles_from_data=False)
    chart_opex.set_categories(cats_opex)
    chart_opex.width = 18
    chart_opex.height = 9.5
    ws_dre.add_chart(chart_opex, "E20")
    
    # Salvar a planilha final
    output_filename = "DRE_Provedor_ISP.xlsx"
    wb.save(output_filename)
    print(f"Planilha {output_filename} gerada com sucesso!")

if __name__ == "__main__":
    generate_dre()
